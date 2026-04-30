from __future__ import annotations

import csv
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
import re
import sys
from typing import Optional

from openpyxl import load_workbook


WORKBOOK_2024 = Path(r"C:\Users\cnarayana2\Downloads\Current Grain2024D (1).xlsm")
WORKBOOK_2025 = Path(r"C:\Users\cnarayana2\OneDrive - University of Nebraska-Lincoln\Grain_SW\Grain2025D_example_data.xlsm")
CROP_CSV_PATH = Path(r"C:\Users\cnarayana2\Downloads\crop.csv")
DELIVERY_LOCATION_CSV_PATH = Path(r"C:\Users\cnarayana2\Downloads\delivery location.csv")
STORAGE_LOCATION_CSV_PATH = Path(r"C:\Users\cnarayana2\Downloads\storage loc table.csv")
OUTPUT_DIR = Path(r"C:\Users\cnarayana2\OneDrive - University of Nebraska-Lincoln\Grain_SW\GMS\outputs")


@dataclass
class DeliveryRow:
    load_ref: str
    ticket_num: int
    delivery_date: str
    crop_code: str
    delivery_to_code: str
    storage_code: str
    mc: Optional[float]
    gross_weight: Optional[float]
    tare_weight: Optional[float]
    bushels: Optional[float]
    delivery_status: Optional[str]
    price: Optional[float]
    gross_sale: Optional[float]
    sale_discounts: Optional[float]
    comments: Optional[str]


def sql_quote(value: Optional[str]) -> str:
    if value is None:
        return "NULL"
    return "'" + value.replace("\\", "\\\\").replace("'", "''") + "'"


def sql_num(value: Optional[float | int], decimals: Optional[int] = None) -> str:
    if value is None:
        return "NULL"
    if isinstance(value, int):
        return str(value)
    num = float(value)
    if decimals is not None:
        num = round(num, decimals)
        return f"{num:.{decimals}f}"
    return f"{num:.4f}".rstrip("0").rstrip(".")


def normalize_text(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def parse_ticket_num(load_value: str) -> Optional[int]:
    raw = normalize_text(load_value)
    if not raw:
        return None
    if raw.isdigit():
        return int(raw)
    if re.fullmatch(r"S\d+", raw, flags=re.IGNORECASE):
        return int(raw[1:])
    if re.fullmatch(r"\d+-\d+", raw):
        return int(raw.replace("-", ""))
    return None


def append_comment(load_ref: str, comment: str) -> str:
    load_note = f"LoadRef={load_ref}"
    if comment:
        return f"{load_note}; {comment}"
    return load_note


def as_date_string(value) -> Optional[str]:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if hasattr(value, "isoformat"):
        return value.isoformat()
    return str(value)


def load_csv_set(path: Path, key: str) -> set[str]:
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        return {normalize_text(row[key]) for row in reader}


def read_delivery_rows(workbook_path: Path) -> list[DeliveryRow]:
    wb = load_workbook(workbook_path, data_only=True, keep_vba=True)
    ws = wb["Delivery"]
    rows: list[DeliveryRow] = []

    for r in range(5, ws.max_row + 1):
        values = [ws.cell(r, c).value for c in range(1, 15)]
        if all(v is None for v in values):
            continue

        load_ref = normalize_text(values[0])
        delivery_date = as_date_string(values[1])
        crop_code = normalize_text(values[2]).upper()
        delivery_to_code = normalize_text(values[3])
        storage_code = normalize_text(values[7])
        bushels = values[8]

        if not delivery_date or not crop_code or not delivery_to_code or not storage_code or bushels is None:
            continue

        comment = normalize_text(values[13])
        ticket_num = parse_ticket_num(load_ref)
        if ticket_num is None:
            ticket_num = -(100000 + r)

        comments = comment or None
        if normalize_text(load_ref) != str(ticket_num):
            comments = append_comment(load_ref, comment)

        rows.append(
            DeliveryRow(
                load_ref=load_ref,
                ticket_num=ticket_num,
                delivery_date=delivery_date,
                crop_code=crop_code,
                delivery_to_code=delivery_to_code,
                storage_code=storage_code,
                mc=float(values[4]) if values[4] is not None else None,
                gross_weight=float(values[5]) if values[5] is not None else None,
                tare_weight=float(values[6]) if values[6] is not None else None,
                bushels=float(values[8]) if values[8] is not None else None,
                delivery_status=normalize_text(values[9]) or None,
                price=float(values[10]) if values[10] is not None else None,
                gross_sale=float(values[11]) if values[11] is not None else None,
                sale_discounts=float(values[12]) if values[12] is not None else None,
                comments=comments,
            )
        )

    return rows


def validate_mappings(rows: list[DeliveryRow]) -> None:
    crop_codes = load_csv_set(CROP_CSV_PATH, "Crop_Code")
    delivery_codes = load_csv_set(DELIVERY_LOCATION_CSV_PATH, "DelLoc_code")
    storage_codes = load_csv_set(STORAGE_LOCATION_CSV_PATH, "Bin_Code")

    unresolved_crops = sorted({r.crop_code for r in rows if r.crop_code not in crop_codes})
    unresolved_delivery = sorted({r.delivery_to_code for r in rows if r.delivery_to_code not in delivery_codes and r.delivery_to_code != "ANF"})
    unresolved_storage = sorted({r.storage_code for r in rows if r.storage_code not in storage_codes})

    if unresolved_crops or unresolved_delivery or unresolved_storage:
        parts = []
        if unresolved_crops:
            parts.append(f"crop codes: {unresolved_crops}")
        if unresolved_delivery:
            parts.append(f"delivery locations: {unresolved_delivery}")
        if unresolved_storage:
            parts.append(f"storage locations: {unresolved_storage}")
        raise ValueError("Unresolved mappings found: " + "; ".join(parts))


def build_insert_sql(rows: list[DeliveryRow]) -> str:
    statements: list[str] = []

    statements.append("START TRANSACTION;")
    statements.append(
        "INSERT INTO delivery_location (DelLoc_code, Location)\n"
        "SELECT 'ANF', 'Feedmill Bins (ANF)'\n"
        "WHERE NOT EXISTS (\n"
        "  SELECT 1 FROM delivery_location WHERE DelLoc_code = 'ANF'\n"
        ");"
    )

    for row in rows:
        statements.append(
            "INSERT INTO delivery (\n"
            "  Ticket_Num, Delivery_Date, Crop_ID, DelLoc_ID, StorLoc_ID,\n"
            "  MC, Gross_Weight, Tare_Weight, Bushels,\n"
            "  Delivery_Status, Price, Gross_Sale, Sale_Discounts, Comments\n"
            ")\nVALUES (\n"
            f"  {sql_num(row.ticket_num)}, {sql_quote(row.delivery_date)},\n"
            f"  (SELECT Crop_ID FROM crop WHERE Crop_Code = {sql_quote(row.crop_code)} ORDER BY Crop_ID LIMIT 1),\n"
            f"  (SELECT DelLoc_ID FROM delivery_location WHERE DelLoc_code = {sql_quote(row.delivery_to_code)} ORDER BY DelLoc_ID LIMIT 1),\n"
            f"  (SELECT StorLoc_ID FROM storage_location WHERE Bin_Code = {sql_quote(row.storage_code)} ORDER BY StorLoc_ID LIMIT 1),\n"
            f"  {sql_num(row.mc, 2)}, {sql_num(row.gross_weight, 2)}, {sql_num(row.tare_weight, 2)}, {sql_num(row.bushels, 2)},\n"
            f"  {sql_quote(row.delivery_status)}, {sql_num(row.price, 4)}, {sql_num(row.gross_sale, 2)}, {sql_num(row.sale_discounts, 2)}, {sql_quote(row.comments)}\n"
            ");"
        )

    statements.append("COMMIT;")
    return "\n\n".join(statements) + "\n"


def main() -> None:
    target = (sys.argv[1] if len(sys.argv) > 1 else "2025").strip().lower()

    if target == "2024":
        workbook_path = WORKBOOK_2024
        output_path = OUTPUT_DIR / "delivery_import_2024.sql"
    elif target == "2025":
        workbook_path = WORKBOOK_2025
        output_path = OUTPUT_DIR / "delivery_import_2025.sql"
    else:
        raise SystemExit("Usage: python generate_delivery_import_sql.py [2024|2025]")

    rows = read_delivery_rows(workbook_path)
    validate_mappings(rows)
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    output_path.write_text(build_insert_sql(rows), encoding="utf-8")

    print(f"Workbook: {workbook_path}")
    print(f"Generated {len(rows)} delivery inserts.")
    print(f"SQL file: {output_path}")
    print(f"Rows with synthetic ticket numbers: {sum(1 for r in rows if r.ticket_num < 0)}")


if __name__ == "__main__":
    main()
